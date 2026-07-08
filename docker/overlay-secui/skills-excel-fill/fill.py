#!/usr/bin/env python3
"""fill.py — excel-fill skill の確定性 Excel 記入エンジン。

winclaw 数字人秘书の「テンプレ xlsx にデータを記入する」系タスク(労働者派遣事業
報告書 等の公式様式)を **毎回 openpyxl コードを現写しない** 確定パイプラインへ固定する。
LLM は「どのセルに何を入れるか」の **spec.json だけ** を作り、記入の機微(合併セル →
アンカー解決 / 保存 / **保存後の再読込検証** / 全ファイル all-or-nothing コミット)は
本スクリプトが担う。これにより:
  - 合併セル非アンカーへの直接代入で openpyxl が例外→ save 到達せず空欄成果物、という
    典型事故を根絶(このスクリプトが自動でアンカーへ解決して書く)。
  - 「書けたつもり」を防ぐ: 保存後に再度開いて指定セルが実際に入っているか検証する。
  - 失敗時は **非ゼロ終了**。呼び元(agent)は exit 0 でない限り task.json/完了報告を
    書いてはならない(= 偽の done を根絶)。

サブコマンド:
  inspect  … テンプレ構造(sheet 名 / 寸法 / 合併範囲 / 非空セルの coord+値)を JSON 出力。
             agent が「ラベルの隣のどのセルへ値を入れるか」を組む手がかりにする。
  fill     … spec に従い記入 → 保存 → 検証 → コミット。全成功で exit 0。

使い方:
  python3 fill.py inspect --template 002634889.xlsx [--sheet 1面] [--max-cells 500]
  python3 fill.py fill --spec spec.json          # 複数ファイル(spec.files[])を一括
  python3 fill.py fill --template T --out O --spec spec.json   # 単一(spec.cells[])

exit code: 0=全成功 / 2=spec/構造エラー(保存前に中止) / 3=検証不一致 / 4=openpyxl 不在。
"""
from __future__ import annotations

import argparse
import json
import os
import sys
from typing import Any


def _die(code: int, msg: str) -> None:
    sys.stderr.write(msg.rstrip("\n") + "\n")
    sys.exit(code)


try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.cell import coordinate_to_tuple
except ImportError:  # pragma: no cover
    _die(
        4,
        "[excel-fill] openpyxl 未導入(image に事前導入が必要)。コードを現写しないこと。",
    )


# ---------------------------------------------------------------------------
# セル座標 / 合併範囲の解決
# ---------------------------------------------------------------------------
def _parse_coord(coord: str) -> tuple[int, int]:
    """'E23' → (row=23, col=5)。不正は ValueError。"""
    row, col = coordinate_to_tuple(str(coord).strip().upper())
    if row < 1 or col < 1:
        raise ValueError(f"不正なセル座標: {coord!r}")
    return row, col


def _anchor(ws, row: int, col: int) -> tuple[int, int]:
    """(row,col) が合併範囲内なら範囲左上(アンカー)へ解決する。合併外はそのまま。

    openpyxl は合併範囲の非アンカーセル(MergedCell)への `.value=` を read-only 例外で
    拒否する。公式様式はラベル横の入力欄が合併されている事が多く、ここを吸収しないと
    記入が全滅する(実際の事故根因)。アンカー(左上)に書けば表示は合併セル全体に出る。
    """
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return rng.min_row, rng.min_col
    return row, col


def _norm_eq(expected: Any, got: Any) -> bool:
    """検証用の緩い等価。数値は数値比較(型差 int/float 吸収)、他は strip 文字列比較。"""
    if expected is None:
        return got is None
    if isinstance(expected, bool) or isinstance(got, bool):
        return bool(expected) == bool(got)
    if isinstance(expected, (int, float)) and isinstance(got, (int, float)):
        try:
            return abs(float(expected) - float(got)) < 1e-9
        except (TypeError, ValueError):
            return str(expected) == str(got)
    return str(expected).strip() == str("" if got is None else got).strip()


# ---------------------------------------------------------------------------
# spec 正規化
# ---------------------------------------------------------------------------
def _normalize_cells(raw: Any) -> list[dict]:
    """spec の cells 部を [{sheet,cell,value}] へ正規化。

    受理形:
      - [{"sheet":..,"cell":..,"value":..}, ...]
      - {"<sheet>": {"<cell>": value, ...}, ...}
    """
    out: list[dict] = []
    if isinstance(raw, list):
        for i, e in enumerate(raw):
            if not isinstance(e, dict) or "sheet" not in e or "cell" not in e:
                raise ValueError(f"cells[{i}] は {{sheet,cell,value}} が必要: {e!r}")
            out.append({"sheet": str(e["sheet"]), "cell": str(e["cell"]), "value": e.get("value")})
    elif isinstance(raw, dict):
        for sheet, cells in raw.items():
            if not isinstance(cells, dict):
                raise ValueError(f"cells['{sheet}'] は {{cell: value}} が必要")
            for cell, value in cells.items():
                out.append({"sheet": str(sheet), "cell": str(cell), "value": value})
    else:
        raise ValueError("cells は配列 or {sheet:{cell:value}} 辞書")
    if not out:
        raise ValueError("cells が空(記入対象が 0 件)")
    return out


def _load_spec(spec_path: str, template: str | None, out: str | None) -> list[dict]:
    """spec.json を読み、記入ジョブ [{template,out,cells:[{sheet,cell,value}]}] を返す。"""
    with open(spec_path, encoding="utf-8") as f:
        spec = json.load(f)

    jobs: list[dict] = []
    if isinstance(spec, dict) and "files" in spec:
        files = spec["files"]
        if not isinstance(files, list) or not files:
            raise ValueError("spec.files は非空配列である必要")
        for i, fj in enumerate(files):
            if not isinstance(fj, dict) or "template" not in fj or "out" not in fj:
                raise ValueError(f"files[{i}] は {{template,out,cells}} が必要")
            jobs.append({
                "template": str(fj["template"]),
                "out": str(fj["out"]),
                "cells": _normalize_cells(fj.get("cells")),
            })
    else:
        # 単一ファイル形。--template/--out 必須。spec は {cells:[...]} or 直接 cells。
        if not template or not out:
            raise ValueError("単一ファイル形は --template と --out が必須")
        cells_raw = spec.get("cells") if isinstance(spec, dict) else spec
        jobs.append({"template": template, "out": out, "cells": _normalize_cells(cells_raw)})
    return jobs


# ---------------------------------------------------------------------------
# inspect
# ---------------------------------------------------------------------------
def cmd_inspect(args: argparse.Namespace) -> None:
    if not os.path.isfile(args.template):
        _die(2, f"[excel-fill] テンプレが見つかりません: {args.template}")
    wb = openpyxl.load_workbook(args.template, data_only=False)
    result: dict = {"file": args.template, "sheets": []}
    for ws in wb.worksheets:
        if args.sheet and ws.title != args.sheet:
            continue
        merged = [str(r) for r in ws.merged_cells.ranges]
        cells = []
        truncated = False
        for row in ws.iter_rows():
            for c in row:
                if c.value is None:
                    continue
                if len(cells) >= args.max_cells:
                    truncated = True
                    break
                v = c.value
                sv = str(v)
                cells.append({
                    "coord": f"{get_column_letter(c.column)}{c.row}",
                    "value": (sv[:80] + "…") if len(sv) > 80 else sv,
                })
            if truncated:
                break
        result["sheets"].append({
            "name": ws.title,
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "merged": merged,
            "cells": cells,
            "cells_truncated": truncated,
        })
    print(json.dumps(result, ensure_ascii=False, indent=2))


# ---------------------------------------------------------------------------
# fill
# ---------------------------------------------------------------------------
def _fill_one(job: dict) -> tuple[Any, list[dict], list[str]]:
    """1 ファイルを記入(まだ保存しない)。(wb, 検証ジョブ, エラー) を返す。

    検証ジョブ = [{sheet, row, col, value}](アンカー解決後の実書込先)。
    """
    errors: list[str] = []
    template, out = job["template"], job["out"]
    if not os.path.isfile(template):
        return None, [], [f"テンプレが見つかりません: {template}"]
    try:
        wb = openpyxl.load_workbook(template, data_only=False)
    except Exception as e:  # noqa: BLE001
        return None, [], [f"テンプレ読込失敗 {template}: {type(e).__name__}: {e}"]

    verify: list[dict] = []
    for e in job["cells"]:
        sheet = e["sheet"]
        if sheet not in wb.sheetnames:
            errors.append(f"[{out}] sheet '{sheet}' が無い(実在: {wb.sheetnames})")
            continue
        ws = wb[sheet]
        try:
            row, col = _parse_coord(e["cell"])
        except ValueError as ve:
            errors.append(f"[{out}] {ve}")
            continue
        a_row, a_col = _anchor(ws, row, col)
        try:
            ws.cell(row=a_row, column=a_col).value = e["value"]
        except Exception as we:  # noqa: BLE001 — 想定外でも spec 単位で握る
            errors.append(
                f"[{out}] {sheet}!{e['cell']}(→{get_column_letter(a_col)}{a_row}) 書込失敗: "
                f"{type(we).__name__}: {we}"
            )
            continue
        verify.append({"sheet": sheet, "row": a_row, "col": a_col, "value": e["value"]})

    # Excel で開いた時に数式(入力用シート→表示シート等)を再計算させる。
    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:  # noqa: BLE001 — openpyxl バージョン差は無視
        pass
    return wb, verify, errors


def _verify_one(out_part: str, verify: list[dict]) -> list[str]:
    """保存済ファイルを再読込し、指定セルが実際に入っているか検証。不一致を返す。"""
    problems: list[str] = []
    try:
        wb = openpyxl.load_workbook(out_part, data_only=False)
    except Exception as e:  # noqa: BLE001
        return [f"再読込失敗 {out_part}: {type(e).__name__}: {e}"]
    for v in verify:
        got = wb[v["sheet"]].cell(row=v["row"], column=v["col"]).value
        if not _norm_eq(v["value"], got):
            problems.append(
                f"{v['sheet']}!{get_column_letter(v['col'])}{v['row']} "
                f"期待={v['value']!r} 実際={got!r}"
            )
    return problems


def cmd_fill(args: argparse.Namespace) -> None:
    jobs = _load_spec(args.spec, args.template, args.out)

    # 1) 全ファイルを記入(まだ保存しない)。1 つでも spec/構造エラーなら保存前に中止。
    prepared: list[dict] = []
    all_errors: list[str] = []
    for job in jobs:
        wb, verify, errors = _fill_one(job)
        all_errors.extend(errors)
        if wb is not None and not errors:
            prepared.append({"job": job, "wb": wb, "verify": verify})
    if all_errors:
        _die(
            2,
            "[excel-fill] spec/構造エラー(保存せず中止・成果物は作らない):\n  - "
            + "\n  - ".join(all_errors)
            + "\n→ inspect でシート名/セルを確認し spec を直すこと。コードは現写しない。",
        )

    # 2) 全ファイルを .part へ保存。
    part_paths: list[str] = []
    try:
        for p in prepared:
            out = p["job"]["out"]
            os.makedirs(os.path.dirname(os.path.abspath(out)) or ".", exist_ok=True)
            # ★一時ファイルは **元の拡張子を保つ**(openpyxl は拡張子で判定し `.part` を拒否する)。
            root, ext = os.path.splitext(out)
            part = f"{root}.part{ext or '.xlsx'}"
            p["wb"].save(part)
            p["part"] = part
            part_paths.append(part)
    except Exception as e:  # noqa: BLE001
        for part in part_paths:
            try:
                os.remove(part)
            except OSError:
                pass
        _die(2, f"[excel-fill] 保存失敗: {type(e).__name__}: {e}")

    # 3) 検証(再読込)。1 つでも不一致なら全 .part を破棄(部分成果物を残さない)。
    verify_problems: list[str] = []
    for p in prepared:
        probs = _verify_one(p["part"], p["verify"])
        if probs:
            verify_problems.append(f"[{p['job']['out']}]")
            verify_problems.extend("  " + x for x in probs)
    if verify_problems:
        for part in part_paths:
            try:
                os.remove(part)
            except OSError:
                pass
        _die(
            3,
            "[excel-fill] 検証不一致(成果物は作らない):\n  "
            + "\n  ".join(verify_problems)
            + "\n→ 値の型(数値は数字/文字は文字)やセル指定を見直すこと。",
        )

    # 4) 全成功 → コミット(原子 rename)。
    for p in prepared:
        os.replace(p["part"], p["job"]["out"])

    total_cells = sum(len(p["verify"]) for p in prepared)
    print(
        f"[excel-fill] OK: {len(prepared)} ファイル / {total_cells} セル記入・検証済 → "
        + ", ".join(os.path.basename(p["job"]["out"]) for p in prepared)
    )


def main() -> None:
    ap = argparse.ArgumentParser(description="excel-fill 確定性 Excel 記入エンジン")
    sub = ap.add_subparsers(dest="cmd", required=True)

    ap_i = sub.add_parser("inspect", help="テンプレ構造を JSON 出力")
    ap_i.add_argument("--template", required=True, help="テンプレ xlsx パス")
    ap_i.add_argument("--sheet", default=None, help="特定シートのみ(既定=全シート)")
    ap_i.add_argument("--max-cells", type=int, default=500, help="シート毎の非空セル上限")
    ap_i.set_defaults(func=cmd_inspect)

    ap_f = sub.add_parser("fill", help="spec に従い記入・保存・検証")
    ap_f.add_argument("--spec", required=True, help="spec.json パス")
    ap_f.add_argument("--template", default=None, help="単一ファイル形の元テンプレ")
    ap_f.add_argument("--out", default=None, help="単一ファイル形の出力先")
    ap_f.set_defaults(func=cmd_fill)

    args = ap.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
